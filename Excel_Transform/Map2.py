import csv
from openpyxl import load_workbook

workbook = load_workbook(filename = 'Map.xlsx')
sheet = workbook.active

criteria_list = []
for row in sheet.iter_rows(min_row = 5, values_only=True):
    if row[0] is not None:
        criteria_list.append(row)

course_id = []
for row in sheet.iter_rows(min_row=3,max_row=3,values_only=True):
    for value in row:
        course_id.append(value)

print(criteria_list)
print(course_id)

cid_criteria = []
for item in range(len(criteria_list)):
    print(criteria_list[item])
    #cid_criteria = []
    crew_name = criteria_list[item][0]
    site = criteria_list[item][1]
    plant_total = criteria_list[item][2]
    per_crew = criteria_list[item][3]

    if 'base' in crew_name.lower():
        crew_type = 'BASE'
    else:
        crew_type = 'NON BASE'

    for cid, criteria in zip(course_id, criteria_list[item]):
        if (cid is not None and cid.lower() != 'courseid'
        and cid.lower() !=  'ops_score' and cid.lower() != 'non_ops_score'):
            if criteria is None:
                criteria = 0
            print(f'{cid} criteria is {criteria}')
            #cid_criteria.append((cid,criteria))
            cid_tuple = (crew_name, cid, criteria, site, plant_total, per_crew, crew_type)
            cid_criteria.append(cid_tuple)

with open('cid_criteria.csv', 'w', newline='') as filename:
    c = csv.writer(filename)
    for row in cid_criteria:
        c.writerow(row)
