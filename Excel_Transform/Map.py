import csv

from openpyxl import load_workbook
workbook = load_workbook(filename='Map.xlsx')
sheet = workbook.active

# Identify the index for every skill
craft = []
for row in sheet.iter_rows(min_row=1,max_row=1,values_only=True):
   ops_index = row.index('OPS')
   nonops_index= row.index('NON_OPS')
   for value in row:
       craft.append(value)

#print(craft)
# These set of for loops should be repeated for every type of index
for index in range(ops_index+1, nonops_index):
    craft[index] = craft[ops_index]

for index in range(nonops_index+ 1, len(craft)):
    craft[index] = craft[nonops_index]

#print(craft)

course_name = []
for row in sheet.iter_rows(min_row=2,max_row=2,values_only=True):
    #print(row)
    for value in row:
        course_name.append(value)


course_id = []
for row in sheet.iter_rows(min_row=3,max_row=3,values_only=True):
    #print(row)
    for value in row:
        course_id.append(value)


#zipped= zip(course_name, course_id, craft)


# for name, cid, skill in zip(course_name, course_id, craft):
#     if name is not None and name != 'CourseName':
#         if skill == 'OPS':
#             craft_code = 'O'
#         elif skill == 'NON_OPS':
#             craft_code = 'N'
#         print(f'{name} with id {cid} belongs to {skill}  \
#              which has the code {craft_code}')


with open('course_list.csv', 'w', newline ='') as file:
    c = csv.writer(file)
    for name,cid, skill in zip(course_name, course_id, craft):
        if name is not None and name != 'CourseName':
            if skill == 'OPS':
                craft_code = 'O'
            elif skill == 'NON_OPS':
                craft_code = 'N'
            print(f'{name} with id {cid} belongs to {skill}  which has the code {craft_code}')
            csv_tuple = (name,cid,skill,craft_code)
            c.writerow(csv_tuple)
